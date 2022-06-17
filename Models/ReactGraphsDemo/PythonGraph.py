import sys
from flask import jsonify
from openpyxl import load_workbook


def getFood_Sales():
    try:
        print("test")
        book = load_workbook('C:/Users/Azees Fetuga/Desktop/Projects/Global API/Models/ReactGraphsDemo/Excels/Food.xlsx',data_only=True)
        print("test2")
        ws = book.active
        print("test3")
        sales = []
        print("test4")
        for row in ws.iter_rows(min_row=2, min_col=1,values_only = True):  
            food_sales = {
                "date": row[0],
                "region": row[1],
                "city": row[2],
                "category": row[3],
                "product": row[4],
                "quantity": row[5],
                "unit_price": row[6],
                "total_price": row[7]
            }
            sales.append(food_sales)
        return jsonify ({"status": True, "Data": sales})
    except:
        print(sys.exc_info()[0])
        return jsonify({"status": False, "error": "data could not be retrieved"})

